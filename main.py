import openpyxl
import os 
import ipdb

def GetFileList(dir, fileList):
    newDir = dir
    if os.path.isfile(dir):         # 如果是文件则添加进 fileList
        fileList.append(dir)
    elif os.path.isdir(dir):
        for s in os.listdir(dir):   # 如果是文件夹
            newDir = os.path.join(dir, s)
            GetFileList(newDir, fileList)
    return fileList
files =[]
f = []
GetFileList('d:/sh/getDate/data/',files)
GetFileList('d:/sh/getDate/todo/',f)
f=f[0]


# 获取待查询的列表
nlist=[]
cc = 1
b= openpyxl.load_workbook(f)
for x in b.get_sheet_names():
    max_row=b[x].max_row
    for y in range(2,max_row+1):
        nlist.append( b[x].cell(row=y,column=cc).value )

# 获取可以查询的列表
data_dic={}
data_cc = 1
name_cc = 1
data_name = '订货日期'
name_name = '订单编号'
for data_f in files:
    # 遍历文件
    b= openpyxl.load_workbook(data_f)

    # 查找名称
    for x in b.get_sheet_names():
        max_row=b[x].max_row
        max_column=b[x].max_column
        print(max_row)
        print(max_column)
        for z in range(1,max_column+1):
            if b[x].cell(row=1,column=z).value == name_name:
                name_cc = z
                break

    # 查找订货日期
    for x in b.get_sheet_names():
        max_row=b[x].max_row
        max_column=b[x].max_column
        for z in range(1,max_column+1):
            print(z)
            if b[x].cell(row=1,column=z).value == data_name:
                data_cc = z
                break
    for x in b.get_sheet_names():
        for y in range(2,max_row+1):
            print(y)
            data_dic[b[x].cell(row=y,column=name_cc).value]=b[x].cell(row=y,column=data_cc).value

print(data_dic)
